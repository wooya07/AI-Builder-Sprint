from html import unescape
from io import BytesIO
import json
from pathlib import Path
import re

from openpyxl import load_workbook


CATALOG_JSON_PATH = Path(__file__).resolve().parent.parent / "data" / "course_catalog.json"
DAY_CODES = {"월": "MON", "화": "TUE", "수": "WED", "목": "THU", "금": "FRI"}

FIELD_ALIASES = {
    "grade": ("\ud559\ub144", "\uad8c\uc7a5\ud559\ub144"),
    "subject_name": ("\uad50\uacfc\ubaa9\uba85(\ubbf8\ud655\uc815\uad6c\ubd84)", "\uad50\uacfc\ubaa9\uba85", "\uacfc\ubaa9\uba85"),
    "course_code": ("\uad50\uacfc\ubaa9\ubc88\ud638", "\uacfc\ubaa9\ucf54\ub4dc", "\uad50\uacfc\ubaa9\ucf54\ub4dc"),
    "section": ("\ubd84\ubc18",),
    "credits": ("\ud559\uc810",),
    "instructor": ("\uad50\uc218\uba85", "\ub2f4\ub2f9\uad50\uc218", "\uad50\uc218"),
    "schedule_text": ("\uc2dc\uac04/\uac15\uc758\uc2e4", "\uc2dc\uac04", "\uc218\uc5c5\uc2dc\uac04", "\uac15\uc758\uc2dc\uac04"),
    "department": ("\uac1c\uc124\ud559\uacfc", "\uac1c\uc124\ubd80\uc11c"),
}
OPTIONAL_FIELDS = {"\uc774\uc218\uad6c\ubd84": "subject_type", "\uad50\uacfc\ubaa9\uad6c\ubd84": "subject_type", "\uad50\uacfc\uad6c\ubd84": "subject_type"}
MEETING_PATTERN = re.compile(
    r"(?P<days>[\uC6D4\uD654\uC218\uBAA9\uAE08](?:\s*[,/·]\s*[\uC6D4\uD654\uC218\uBAA9\uAE08])*)(?:\uc694\uc77c)?\s*"
    r"(?P<hour>\d{1,2})\s*:\s*(?P<minute>\d{2})\s*"
    r"\(\s*(?P<duration>\d+)\s*(?:\ubd84)?\)\s*"
    r"(?P<building>[A-Za-z0-9\uAC00-\uD7A3]+)\s*-\s*(?P<room>[A-Za-z0-9-]+)"
)
RANGE_MEETING_PATTERN = re.compile(
    r"(?P<days>[\uC6D4\uD654\uC218\uBAA9\uAE08](?:\s*[,/·]\s*[\uC6D4\uD654\uC218\uBAA9\uAE08])*)(?:\uc694\uc77c)?\s*"
    r"(?P<start_hour>\d{1,2})\s*:\s*(?P<start_minute>\d{2})\s*[-~]\s*"
    r"(?P<end_hour>\d{1,2})\s*:\s*(?P<end_minute>\d{2})\s*"
    r"(?P<building>[A-Za-z0-9\uAC00-\uD7A3]+)\s*-\s*(?P<room>[A-Za-z0-9-]+)"
)


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def parse_schedule(schedule_text: str) -> list[dict[str, str | int]]:
    """Parse weekday, time, duration (minutes), building, and room values."""
    normalized = unescape(schedule_text).translate(str.maketrans({"（": "(", "）": ")", "：": ":", "－": "-", "～": "~"}))
    normalized = re.sub(r"<br\s*/?>", "\n", normalized, flags=re.IGNORECASE)
    meetings: list[dict[str, str | int]] = []
    for match in MEETING_PATTERN.finditer(normalized):
        hour, minute, duration = int(match["hour"]), int(match["minute"]), int(match["duration"])
        start_minutes = hour * 60 + minute
        if hour <= 23 and minute <= 59 and duration > 0 and start_minutes + duration <= 24 * 60:
            for day in _split_days(match["days"]):
                meetings.append({
                    "day": day, "start_minutes": start_minutes,
                    "duration_minutes": duration, "building": match["building"], "room": match["room"],
                })
    for match in RANGE_MEETING_PATTERN.finditer(normalized):
        start_minutes = int(match["start_hour"]) * 60 + int(match["start_minute"])
        end_minutes = int(match["end_hour"]) * 60 + int(match["end_minute"])
        if start_minutes < end_minutes <= 24 * 60:
            for day in _split_days(match["days"]):
                meetings.append({
                    "day": day, "start_minutes": start_minutes,
                    "duration_minutes": end_minutes - start_minutes,
                    "building": match["building"], "room": match["room"],
                })
    unique_meetings: dict[tuple[str, int, str, str], dict[str, str | int]] = {}
    for meeting in meetings:
        key = (str(meeting["day"]), int(meeting["start_minutes"]), str(meeting["building"]), str(meeting["room"]))
        unique_meetings[key] = meeting
    return list(unique_meetings.values())


def _split_days(value: str) -> list[str]:
    return re.findall(r"[\uC6D4\uD654\uC218\uBAA9\uAE08]", value)


def read_catalog_excel(content: bytes) -> tuple[list[dict[str, object]], list[int]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as error:
        raise ValueError("\uc5d1\uc140 \ud30c\uc77c\uc774 \ube44\uc5b4 \uc788\uc2b5\ub2c8\ub2e4.") from error

    indexed_headers = {normalize_header(value): index for index, value in enumerate(header_row)}
    field_indices = {
        field: next((indexed_headers[normalize_header(alias)] for alias in aliases if normalize_header(alias) in indexed_headers), None)
        for field, aliases in FIELD_ALIASES.items()
    }
    missing = [FIELD_ALIASES[field][0] for field, index in field_indices.items() if index is None]
    if missing:
        raise ValueError("\ud544\uc218 \uc5f4\uc774 \uc5c6\uc2b5\ub2c8\ub2e4: " + ", ".join(missing))

    records: list[dict[str, object]] = []
    skipped_rows: list[int] = []
    for row_number, row in enumerate(rows, start=2):
        def value(field: str) -> str:
            index = field_indices[field]
            return str(row[index] or "").strip() if index is not None and index < len(row) else ""

        try:
            credits = int(float(value("credits")))
            schedule_text = value("schedule_text")
            record: dict[str, object] = {
                "grade": value("grade"), "subject_name": value("subject_name"),
                "course_code": value("course_code"), "section": value("section"), "credits": credits,
                "instructor": value("instructor"), "schedule_text": schedule_text,
                "department": value("department"), "subject_type": "", "meetings": parse_schedule(schedule_text),
            }
            for header, target in OPTIONAL_FIELDS.items():
                if normalize_header(header) in indexed_headers:
                    record[target] = str(row[indexed_headers[normalize_header(header)]] or "").strip()
                    break
            if not all(str(record[field]).strip() for field in FIELD_ALIASES if field != "credits") or credits < 1 or not record["meetings"]:
                raise ValueError
            records.append(record)
        except (ValueError, TypeError, IndexError):
            skipped_rows.append(row_number)
    if not records:
        raise ValueError("\uc800\uc7a5\ud560 \uc218 \uc788\ub294 \uac15\uc758\uac00 \uc5c6\uc2b5\ub2c8\ub2e4. \uc2dc\uac04 \uc5f4\uc740 '\uc6d4 15:00(75) 306-310' \ud615\uc2dd\uc73c\ub85c \uc791\uc131\ud574 \uc8fc\uc138\uc694.")
    return records, skipped_rows


def save_catalog_json(records: list[dict[str, object]]) -> tuple[int, int]:
    """Append only previously unseen class groups to the timetable JSON catalog."""
    existing_payload = _load_catalog_payload()
    existing_classes = existing_payload["classes"]
    existing_group_ids = {
        str(course["classGroupId"])
        for course in existing_classes
        if isinstance(course, dict) and course.get("classGroupId")
    }

    classes: list[dict[str, object]] = []
    imported_count = 0
    duplicate_count = 0
    for record in records:
        class_group_id = f"{record['course_code']}-{record['section']}"
        if class_group_id in existing_group_ids:
            duplicate_count += 1
            continue

        imported_count += 1
        existing_group_ids.add(class_group_id)
        for meeting in record["meetings"]:
            start_minutes = int(meeting["start_minutes"])
            end_minutes = start_minutes + int(meeting["duration_minutes"])
            classes.append({
                "courseId": str(record["course_code"]),
                "classGroupId": class_group_id,
                "courseName": str(record["subject_name"]),
                "day": DAY_CODES[str(meeting["day"])],
                "startTime": _format_time(start_minutes),
                "endTime": _format_time(end_minutes),
                "grade": str(record["grade"]),
                "courseType": str(record["subject_type"]),
                "section": str(record["section"]),
                "credits": int(record["credits"]),
                "instructor": str(record["instructor"]),
                "department": _normalise_department(str(record["department"])),
                "location": {
                    "building": str(meeting["building"]),
                    "room": str(meeting["room"]),
                },
            })

    if not classes:
        return imported_count, duplicate_count

    CATALOG_JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = CATALOG_JSON_PATH.with_suffix(".json.tmp")
    temporary_path.write_text(
        json.dumps(
            {"timezone": existing_payload["timezone"], "classes": [*existing_classes, *classes]},
            ensure_ascii=False,
            indent=2,
        ) + "\n",
        encoding="utf-8",
    )
    temporary_path.replace(CATALOG_JSON_PATH)
    return imported_count, duplicate_count


def _load_catalog_payload() -> dict[str, object]:
    if not CATALOG_JSON_PATH.exists():
        return {"timezone": "Asia/Seoul", "classes": []}

    try:
        payload = json.loads(CATALOG_JSON_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise ValueError("기존 수강편람 JSON을 읽을 수 없습니다.") from error

    classes = payload.get("classes")
    if not isinstance(classes, list):
        raise ValueError("기존 수강편람 JSON 형식이 올바르지 않습니다.")
    timezone = payload.get("timezone")
    return {"timezone": timezone if isinstance(timezone, str) else "Asia/Seoul", "classes": classes}


def _format_time(minutes: int) -> str:
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def _normalise_department(value: str) -> str:
    """Remove the contact line that follows the department name in the source."""
    text = unescape(value)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    return next((" ".join(line.split()) for line in text.splitlines() if line.strip()), "")


def import_catalog(content: bytes) -> tuple[int, int, list[int]]:
    records, skipped_rows = read_catalog_excel(content)
    imported_count, duplicate_count = save_catalog_json(records)
    return imported_count, duplicate_count, skipped_rows
